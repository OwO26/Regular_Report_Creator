import streamlit as st
import pandas as pd
import os
import tempfile

st.set_page_config(page_title="Regular Report Creator", layout="wide")
st.title("📊 Semi-automatic Regular Report Creator")

st.markdown("""
### Pre-work Instructions

Before updating the files, please:

1. Download the latest files from **Acolaid** using the following conditions:  
   - `StatClass` = Q1 to Q6 (separately)  
   - `Decision Date` is null

2. Manually add a column named **Meeting Date** and input the relevant data.

3. Ensure files are named as `Q1.csv`, `Q2.csv`, etc.

4. Only **CSV** files are accepted.

---
""")


uploaded_files = st.file_uploader("Please select and upload multiple CSV files", type=["csv"], accept_multiple_files=True)

file_paths = {}
if uploaded_files:
    st.success(f"✅ Total uploaded {len(uploaded_files)} file(s)")
    st.write("file name：", [file.name for file in uploaded_files])
    
    for uploaded_file in uploaded_files:
        filename = uploaded_file.name
        if filename.upper().startswith("Q") and filename.upper().endswith(".CSV"):
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
            temp_file.write(uploaded_file.read())
            temp_file.close()
            file_paths[filename[:-4]] = temp_file.name

    if st.button("Create a Regular Report"):

        st.info("Script Running")

        import pandas as pd
        from datetime import datetime
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill

        
        
        dataframes = {}
        for name, path in file_paths.items():
            df = pd.read_csv(path, encoding="ISO-8859-1")
            dataframes[name] = df
        
        merged_df = pd.concat(dataframes.values(), ignore_index=True)
        today = datetime.today()
        merged_df['Reg Date'] = pd.to_datetime(merged_df['Reg Date'], errors='coerce')
        merged_df['No. of weeks in system'] = merged_df['Reg Date'].apply(
            lambda x: (today - x).days // 7 if pd.notnull(x) else 'N/A'
        )
        merged_df['Expiry Date'] = pd.to_datetime(merged_df['Expiry Date'], errors='coerce')
        merged_df['No. of weeks past expiry date'] = merged_df['Expiry Date'].apply(
            lambda x: (today - x).days // 7 if pd.notnull(x) else 'N/A'
        )
        merged_df['Meeting Date'] = pd.to_datetime(merged_df['Meeting Date'], errors='coerce')
        merged_df['No. of weeks past meeting date'] = merged_df['Meeting Date'].apply(
            lambda x: (today - x).days // 7 if pd.notnull(x) else 'N/A'
        )
        
        filtered_df = merged_df[merged_df['App Type'] != 'PAS'].copy()

        # 删除 Validation Code 为 "INV" 的行
        if "Validation Code" in filtered_df.columns:
            filtered_df = filtered_df[filtered_df["Validation Code"] != "INV"]
        
        
        desired_columns = [
            "Application Number", "Application Address", "Officer", "Reception Date", "Reg Date",
            "No. of weeks in system", "Expiry Date", "No. of weeks past expiry date", "Meeting Date", 
            "No. of weeks past meeting date", "PPA",
            "App Type", "Finalised Decision Level",
            "Agent Name", "Applicant Name", "Proposal"
        ]
       
        column_renames = {
            "CaseFullRef": "Application Number",
            "Application Address": "Application Address",
            "Officer": "Officer",
            "Reception Date": "Reception Date",
            "Reg Date": "Reg Date",
            "No. of weeks in system": "No. of weeks in system",
            "Expiry Date": "Expiry Date",
            "No. of weeks past expiry date": "No. of weeks past expiry date",
            "Meeting Date": "Meeting Date",
            "No. of weeks past meeting date": "No. of weeks past meeting date",
            "PPA.1": "PPA",
            "App Type": "App Type",
            "Decision Level": "Finalised Decision Level",
            "Agent Name": "Agent Name",
            "Applicant Name": "Applicant Name",
            "Proposal": "Proposal",
        }
        processed_df = filtered_df.rename(columns=column_renames)
        for col in desired_columns:
            if col not in processed_df.columns:
                processed_df[col] = ""
        processed_df = processed_df[desired_columns]
        
        date_format = "%d %b %Y"
        for col in ['Reception Date', 'Reg Date', 'Expiry Date', 'Meeting Date']:
            try:
                processed_df[col] = pd.to_datetime(
                    processed_df[col].astype(str),
                    errors='coerce',
                    dayfirst=True
                ).dt.strftime(date_format).fillna("")
            except Exception as e:
                st.warning(f"⚠️ Date coloumn {col} Unavailable：{e}")

        try:
            processed_df['Reception Date (parsed)'] = pd.to_datetime(
                processed_df['Reception Date'].astype(str),
                format=date_format,
                errors='coerce'
            )
            processed_df = processed_df.sort_values(by='Reception Date (parsed)', ascending=True).drop(columns=['Reception Date (parsed)'])
        except Exception as e:
            st.warning(f"⚠️ Reception Date Order Fail：{e}")

        output_path = "Final_Planning_Table.xlsx"
        processed_df.to_excel(output_path, index=False, engine='openpyxl')
        
        wb = load_workbook(output_path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        proposal_col_index = header.index("Proposal") + 1
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in ws.iter_rows(min_row=2, min_col=proposal_col_index, max_col=proposal_col_index):
            cell = row[0]
            if isinstance(cell.value, str):
                value = cell.value.lower()
                contains_target = any(keyword in value for keyword in ['commercial', 'student', 'student accommodation'])
                lacks_dwell = not any(word in value for word in ['dwell', 'dwelling', 'resident', 'residential', 'c3'])
                if contains_target and lacks_dwell:
                    cell.fill = yellow_fill
        

        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        target_columns = [
            "No. of weeks in system",
            "No. of weeks past expiry date",
            "No. of week past meeting date"
        ]
        for col_name in target_columns:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                values_with_cells = []
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    cell = row[0]
                    try:
                        value = float(cell.value)
                        values_with_cells.append((value, cell))
                    except:
                        continue
                top_cells = sorted(values_with_cells, key=lambda x: x[0], reverse=True)[:10]
                for _, cell in top_cells:
                    cell.fill = light_red_fill
        


        # 插入注释行作为第一行
        ws.insert_rows(1)

        footer_cell = ws.cell(row=1, column=1)
        footer_cell.value = f"This report covers all live applications received up to {datetime.today().strftime('%d %b %Y')}. The yellow highlight suggests applications that may be for commercial or student use only, based on a preliminary keyword search. The red highlight marks the 10 applications with the longest processing times."
        footer_cell.alignment = Alignment(horizontal="center", vertical="center")
        footer_cell.font = Font(bold=True, italic=True)
        footer_cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # 合并注释单元格（横跨所有列）
        last_col = ws.max_column
        merge_range = f"A1:{get_column_letter(last_col)}1"
        ws.merge_cells(merge_range)

        # 删除旧表格（如果存在）
        if ws.tables:
            for tbl in list(ws.tables.values()):
                del ws.tables[tbl.name]

        # 添加新格式化表格，从第2行开始（因为第1行是注释）
        max_row = ws.max_row
        max_col = ws.max_column
        table_range = f"A2:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName="PlanningDataTable", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleLight12",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # 保存最终文件
        wb.save(output_path)
        print(f"✅ Excel file have been saved：{output_path}")

        
        

        st.success("✅ Data processing completed!")

        today_str = datetime.today().strftime("%d_%b_%Y")
        output_filename = f"Regular_Report_on_Pending_Application_{today_str}.xlsx"

        if os.path.exists(output_path):
           with open(output_path, "rb") as f:
              st.download_button(
                 "Click here to download the result file",
                 f,
                 file_name=output_filename
              )
